import urllib2
import time
import csv
import sys
import os
import re
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webelement import WebElement
from _io import open
from _ast import Str
from locale import str
from test.test_iterlen import len
from pip._vendor.distlib.compat import raw_input
from mimify import File
from openpyxl import load_workbook

data = open(os.getcwd() + "/wordOutput.doc", 'w')
wb = load_workbook(filename = 'excl.xlsx')
sheet = wb.worksheets[0]
row_count = sheet.max_row
column_count = 2
count = 1

print "Please wait..."

while count < row_count:
    if sheet['A' + str(count)].value:
        data.write(sheet['A' + str(count)].value + unicode(chr(12)))
    if sheet['B' + str(count)].value:
        data.write(sheet['B' + str(count)].value + unicode(chr(12)))
    count+=1
    

# Generates word file
data.close()

print "Finished document generation"